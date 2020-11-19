import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Font, Side
from openpyxl.styles.colors import YELLOW


# v 0.0.1


class ExcelWriter(object):
    """
    Wrapper Class for writing and appending data into excel files
    """
    def __init__(self, filename, headers):
        """
        :param filename: string
        :param headers: list of strings
        """
        self.filename = filename + '.xlsx'
        self.headers = headers 
        self.data = []
    
    def add(self, data_set):
        """
        Adds data_set to data after checking the keys coincide with the headers.

        Raises DifferentNumberOfKeys
        Raises BadKey

        :param data_set: dictionary
        :return: None
        """
        if len(data_set.keys()) != len(self.headers):
            raise DifferentNumberOfKeys(str(self.headers), str(data_set.keys()))
        else:
            for key in data_set.keys():
                if key not in self.headers:
                    raise BadKey(key)

        # Take care of NoneType (generates problems with sorting)
        for key in data_set:
            if data_set[key] is None:
                data_set[key] = 'None'

        self.data.append(data_set)

    def order_by(self, header_name):
        """
        Order the data by the values of a specific header.

        Raises NonExistentHeaderName

        :param header_name: string
        :return: None
        """
        if header_name not in self.headers:
            raise NonExistentHeaderName(header_name)

        self.data.sort(key= lambda k: k[header_name])

    def write(self):
        """
        Raises UnexpectedHeadersInFile

        :return: None
        """
        # Chose strategy depending on whether the file already exists.
        try:
            test = open(self.filename, 'r')
        except FileNotFoundError:
            self._create()
        else:
            test.close()
            self._append()

    def _check_file_headers(self, worksheet):
        """
        Traverse the columns at row 1 in the excel file and compare with the stored headers.

        Raises UnexpectedHeadersInFile if the file headers don't coincide with the initialization ones.

        :param worksheet: WorkSheet openpyxl object
        :return: None
        """
        file_headers = []

        # Grab all values on first row until finding an empty column.
        column = 1
        cell = worksheet.cell(1, column)
        while cell.value is not None:
            file_headers.append(cell.value)
            column += 1
            cell = worksheet.cell(1, column)

        # Check
        if not file_headers == self.headers:
            raise UnexpectedHeadersInFile(str(self.headers),  str(file_headers))

    def _append(self):
        """
        Writes data into the first empty row in an existing file.

        :return: None
        """
        # Load excel file
        excel_file = load_workbook(self.filename)
        worksheet = excel_file.active

        # Check that the headers coincide
        self._check_file_headers(worksheet)

        # Look for the first empty row.
        row = 1
        cell = worksheet.cell(row, 1)
        while cell.value is not None:
            row += 1
            cell = worksheet.cell(row, 1)

        # Put data and save
        self._put_data(worksheet, row)
        excel_file.save(self.filename)

    def _create(self):
        """
        Creates a new file, puts the headers then the data.

        :return: None
        """
        # Create excel file
        excel_file = Workbook()
        worksheet = excel_file.active

        # Write headers
        for i in range(0, len(self.headers)):
            cell = worksheet.cell(1, i+1)
            cell.value = self.headers[i]
            cell.fill = PatternFill(fill_type='solid', start_color=YELLOW, end_color=YELLOW)
            cell.font = Font(bold=True)
            side_headers = Side(border_style='medium')
            cell.border = Border(left=side_headers, right=side_headers, top=side_headers, bottom=side_headers)

        # Put data and save
        self._put_data(worksheet, 2)
        excel_file.save(self.filename)
    
    def _put_data(self, worksheet, row):
        """
        Puts self.data into an excel worksheet starting from row.

        :param worksheet: WorkSheet openpyxl object
        :param row: integer, row to start writing from
        :return: None
        """

        # Puts all data in self.data
        for data_set in self.data:
            for i in range(0, len(self.headers)):
                try:
                    cell = worksheet.cell(row, i+1)
                    cell.value = data_set[self.headers[i]]
                    side_cell = Side(border_style='thin')
                    cell.border = Border(left=side_cell, right=side_cell, top=side_cell, bottom=side_cell)
                except openpyxl.utils.exceptions.IllegalCharacterError:
                    print("Illegal character error for header %s:  %s" % (self.headers[i], data_set[self.headers[i]]))

            row += 1

        # Cleans self.data
        self.data = []


# EXCEPTIONS

class ExcelWriterError(Exception):
    pass


class NonExistentHeaderName(ExcelWriterError):
    pass


class BadKey(ExcelWriterError):
    pass


class ExpectedVsGotException(ExcelWriterError):
    def __init__(self, expected, got):
        self.expected = expected
        self.got = got
        Exception.__init__(self, got)

    def __str__(self):
        return 'Expected "%s" , got "%s"' % (self.expected, self.got)


class DifferentNumberOfKeys(ExpectedVsGotException):
    pass


class UnexpectedHeadersInFile(ExpectedVsGotException):
    pass
